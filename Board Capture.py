# -*- coding: utf-8 -*-
"""
Created on Thu Jan 25 10:05:45 2018

@author: Maxwell Ledermann
For questions or assistance, email maxwell.ledermann@gmail.com
"""

import cv2 #install as opencv-python
import numpy as np #install
import math
import datetime
import pickle
import openpyxl as pyxl #install
import keyboard #install
import sys

DEBUG = False
scan = False
blank_scan = np.empty((480,640,3))
cv2.imshow("Scan",blank_scan)
print("3414 Hackbots 2019 Pre-Season Scouting Program")

capture_device = 0
cap = cv2.VideoCapture(capture_device)
ret, __ = cap.read()
if ret == False:
    cap = cv2.VideoCapture(capture_device+1)
    cap = cv2.VideoCapture(capture_device)

try:
    wb = pyxl.load_workbook('scouting.xlsx')
    ws = wb.active
except:
    wb = pyxl.Workbook()
    ws = wb.active
    header = ["Time","Team Number","Match Number", "Alliance Station","Starting Position","Plate Config", \
    "Crossed Baseline","Preload Cube","Second Cube","Teleop Scale","Teleop Switch","Teleop Op Switch", \
    "Teleop Vault", "Climbed","Parked","Lifted One","Lifted Two","Was Lifted"]
    for header_column in header:
        ws.cell(row=1, column=(header.index(header_column)+1)).value = header_column

def find_perimeter(poly):
    perimeter = 0
    poly = np.append(poly,[poly[0]],0)
    for i in range(len(poly)-1):
        x0,y0,x1,y1 = poly[i][0],poly[i][1],poly[i+1][0],poly[i+1][1]
        perimeter += math.hypot((x1-x0),(y1-y0))
    return perimeter
    
def find_corner(target,opposite,corners):
    lowest_distance = math.sqrt((target[0]-opposite[0])**2 + (target[1]-opposite[1])**2)
    for corner in corners:
        if (math.sqrt((corner[0]-target[0])**2 + (corner[1]-target[1])**2)) < lowest_distance:
            point = corner
            lowest_distance = math.sqrt((corner[0]-target[0])**2 + (corner[1]-target[1])**2)
    return(point)

while(True):
    if keyboard.is_pressed('s'):
        while keyboard.is_pressed('s'):
            pass
        capture_device += 1
        cap = cv2.VideoCapture(capture_device)
        ret, __ = cap.read()
        if ret == False:
            capture_device = 0
            cap = cv2.VideoCapture(capture_device)

    team_number, match_number, scan = pickle.load(open("entry data.p", "rb"))
    
    ret, frame = cap.read()
    
    cv2.imshow("Capture",frame)
    if keyboard.is_pressed('ctrl+space') == True:
        team_number = 0
        match_number = 0
        scan = True

    if scan == True:
        scan = False
        pickle.dump((0,0,scan), open('entry data.p','wb'))
        font = cv2.FONT_HERSHEY_SIMPLEX
        cv2.putText(frame,str("Team Number: " + str(team_number)),(5,15),font,.5,(0,0,0),3,cv2.LINE_AA)
        cv2.putText(frame,str("Team Number: " + str(team_number)),(5,15),font,.5,(255,255,255),1,cv2.LINE_AA)
        cv2.putText(frame,str("Match Number: " + str(match_number)),(5,30),font,.5,(0,0,0),3,cv2.LINE_AA)
        cv2.putText(frame,str("Match Number: " +str( match_number)),(5,30),font,.5,(255,255,255),1,cv2.LINE_AA)

        #Establish variables 
        starting_position = ["left","center","right"]
        crossed_baseline = True
        preload_cube = ["switch","scale"]
        second_cube = ["switch","scale"]
        found_scale = 10
        found_switch = 10
        found_op_switch = 10
        found_vault = 10
        climbed = True
        parked = True
        lift_one = True
        lift_two = True
        was_lifted = True
        plate_config = ["LLL","RRR","LRL","RLR"]
        alliance_station = ""
        
        targets_highlighted = frame.copy()
        screenshot = frame.copy()    
        
        screenshot = cv2.cvtColor(screenshot,cv2.COLOR_BGR2GRAY)
        #change ret, screenshot = cv2.threshold(screenshot,--->110<---,255,0)
        #this number to adjust for overall brightness
        ret, screenshot = cv2.threshold(screenshot,110,255,0)
        ret, contours, contour_hierarchy = cv2.findContours(screenshot,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)     
    
        contours_big = []
        array = []
        minimum_perimeter = 100
        for pos in range(len(contours)):
            if (find_perimeter(contours[pos][:,0,:]) > minimum_perimeter):
                contours_big += [contours[pos]]
                array.append(pos)
    
        contour_count = 1
        save_data = True
        for current_contour in contours_big:
            print("Scanning " + str(round(float(float((contour_count)/float(len(contours_big))*100.0)),1)) + "%")
            contour_count += 1
            match_score = 0
            variable = 10000
            points = current_contour[:,0,:]
            max_right = max(current_contour[:,0,0])
            max_left = min(current_contour[:,0,0])
            max_up = min(current_contour[:,0,1])
            max_down = max(current_contour[:,0,1])
            center_x = (max(current_contour[:,0,0])+min(current_contour[:,0,0]))/2
            center_y = (min(current_contour[:,0,1])+max(current_contour[:,0,1]))/2
            match_point_tl = [center_x-variable,center_y-variable]
            match_point_tr = [center_x+variable,center_y-variable]
            match_point_bl = [center_x-variable,center_y+variable]
            match_point_br = [center_x+variable,center_y+variable]
        
                        
            top_left = find_corner(match_point_tl,match_point_br,points)    
            top_right = find_corner(match_point_tr,match_point_bl,points)    
            bottom_left = find_corner(match_point_bl,match_point_tr,points)    
            bottom_right = find_corner(match_point_br,match_point_tl,points)
            object_corners = np.float32([top_left,top_right,bottom_left,bottom_right])
            
            pts2 = np.float32([[0,0],[400,0],[0,400],[400,400]])
            M = cv2.getPerspectiveTransform(object_corners,pts2)
            object_warped = cv2.warpPerspective(frame,M,(400,400))   
            object_warped_gray = cv2.cvtColor(object_warped,cv2.COLOR_BGR2GRAY)
            
            for target_to_match in ((range(1,22) + range(79,85))):
                match_template = cv2.imread('targets/target' + str(target_to_match) + '.png',0)
                method = eval('cv2.TM_CCOEFF_NORMED')
                res = cv2.matchTemplate(object_warped_gray,match_template,method)
                min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
    
                if (max_val > match_score):
                    match_score = max_val
                    matched_pts = [top_left,top_right,bottom_right,bottom_left]
                    matched_pts = np.array(matched_pts, np.int32)
                    matched_pts = matched_pts.reshape((-1,1,2))
            
                if match_score > .7:
                    cv2.polylines(targets_highlighted,[matched_pts],True,(0,255,0),2)
                    cv2.imshow("Scan",targets_highlighted)
                    if cv2.waitKey(1) & 0xFF == 27:
                        save_data = False
                        break
                    if target_to_match == 1:
                        starting_position.remove("left")
                    elif target_to_match == 2:
                        starting_position.remove("center")
                    elif target_to_match == 3:
                        starting_position.remove("right")
                    elif target_to_match == 4:
                        crossed_baseline = False
                    elif target_to_match == 5:
                        preload_cube.remove("switch")
                    elif target_to_match == 6:
                        preload_cube.remove("scale")
                    elif target_to_match == 7:
                        second_cube.remove("switch")
                    elif target_to_match == 8:
                        second_cube.remove("scale")
                    elif target_to_match == 9:
                        found_scale -= 1
                    elif target_to_match == 10:
                        found_switch -= 1
                    elif target_to_match == 11:
                        found_op_switch -= 1
                    elif target_to_match == 12:
                        found_vault -= 1
                    elif target_to_match == 13:
                        climbed = False
                    elif target_to_match == 14:
                        parked = False
                    elif target_to_match == 15:
                        lift_one = False
                    elif target_to_match == 16:
                        lift_two = False
                    elif target_to_match == 17:
                        was_lifted = False
                    elif target_to_match == 18:
                        plate_config.remove("LLL")
                    elif target_to_match == 19:
                        plate_config.remove("RRR")
                    elif target_to_match == 20:
                        plate_config.remove("LRL")
                    elif target_to_match == 21:
                        plate_config.remove("RLR")
                    elif target_to_match == 79:
                        alliance_station = "Red1"
                    elif target_to_match == 80:
                        alliance_station = "Red2"
                    elif target_to_match == 81:
                        alliance_station = "Red3"
                    elif target_to_match == 82:
                        alliance_station = "Blue1"
                    elif target_to_match == 83:
                        alliance_station = "Blue2"
                    elif target_to_match == 84:
                        alliance_station = "Blue3"
                    break
                
            if save_data == False:
                break
                
        if save_data == True:
            print("Scan Complete")
    
            if DEBUG == True:
                cv2.imwrite("found targets highlighted.png",targets_highlighted)
            
            export_data = []
            export_data = [alliance_station,starting_position,plate_config,crossed_baseline,preload_cube,second_cube,found_scale,found_switch,found_op_switch,found_vault, \
            climbed,parked,lift_one,lift_two,was_lifted]
            
            for data_index in range(len(export_data)):
                if isinstance(export_data[data_index],list):
                    if len(export_data[data_index]) == 1:
                        export_data[data_index] = str(export_data[data_index][0])
                    elif len(export_data[data_index]) == 0:
                        export_data[data_index] = "NULL"
                    elif len(export_data[data_index]) > 1:
                        export_data[data_index] = "BAD SCAN"
                else:
                    export_data[data_index] = str(export_data[data_index])
                
            now = datetime.datetime.now()
            export_data.insert(0,now)
            export_data.insert(1,team_number)
            export_data.insert(2,match_number)
            ws.append(export_data)
            
            if DEBUG == True:
                print(export_data)
            
            while(True):
                try:
                    wb.save('scouting.xlsx')
                    wb.save('spreadsheet backups/scouting' + str(match_number) + '.xlsx')
                    break
                except:
                    if raw_input("Please close scouting.xlsx | Press enter when closed") == "exit":
                        sys.exit()
        
            save_name = "Saved scans/match" + str(match_number) + "-team" + str(team_number) + ".png"
            cv2.imwrite(save_name, targets_highlighted)
            cv2.imshow("Scan",targets_highlighted)
            print("Data Saved \n")
            
        elif save_data == False:
            print("Scan canceled, no data saved \n")
            while keyboard.is_pressed(chr(27)):
                pass
            
    if cv2.waitKey(1) & 0xFF == 27:
        break

cap.release()
cv2.destroyAllWindows()
"""
Target to 2019 board translation
SANDSTORM
Cross Baseline - 9
Platform 1 - 10
Platform 2 - 11
ENDGAME
Platform 1 - 12
Platform 2 - 13
Platform 3 - 14
Assist - 15
ROCKET
CARGO
LOW - 16
MID - 17
HIGH - 14
HATCH
